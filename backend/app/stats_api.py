# backend/app/stats_api.py
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display

# ── Arabic font registration ──────────────────────────────
# Search for an Arabic-capable TTF on Windows and Linux
def _find_arabic_font() -> tuple[str | None, str | None]:
    import os
    _here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        # Bundled fonts in app/fonts/ (highest priority)
        (os.path.join(_here, "fonts", "NotoNaskhArabic-Regular.ttf"),
         os.path.join(_here, "fonts", "NotoNaskhArabic-Bold.ttf")),
        (os.path.join(_here, "fonts", "Amiri-Regular.ttf"),
         os.path.join(_here, "fonts", "Amiri-Bold.ttf")),
        # Windows — fonts built into every Windows install that support Arabic
        (r"C:\Windows\Fonts\arial.ttf",       r"C:\Windows\Fonts\arialbd.ttf"),
        (r"C:\Windows\Fonts\times.ttf",        r"C:\Windows\Fonts\timesbd.ttf"),
        (r"C:\Windows\Fonts\tahoma.ttf",       r"C:\Windows\Fonts\tahomabd.ttf"),
        (r"C:\Windows\Fonts\calibri.ttf",      r"C:\Windows\Fonts\calibrib.ttf"),
        # Linux system fonts
        ("/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf",
         "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf"),
        ("/usr/share/fonts/truetype/freefont/FreeSerif.ttf",
         "/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ]
    for reg, bold in candidates:
        if os.path.exists(reg):
            return reg, bold if os.path.exists(bold) else reg
    return None, None

_ARABIC_FONT   = "ArabicFont"
_ARABIC_FONT_B = "ArabicFontBold"
_ARABIC_OK = False

_font_reg, _font_bold = _find_arabic_font()
if _font_reg:
    try:
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT,   _font_reg))
        pdfmetrics.registerFont(TTFont(_ARABIC_FONT_B, _font_bold))
        _ARABIC_OK = True
    except Exception:
        _ARABIC_OK = False


def _ar(text: str) -> str:
    """Reshape + bidi-reorder Arabic text so ReportLab renders it correctly."""
    if not text:
        return ""
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)


def _pdf_font(bold: bool = False) -> str:
    if _ARABIC_OK:
        return _ARABIC_FONT_B if bold else _ARABIC_FONT
    return "Helvetica-Bold" if bold else "Helvetica"

from .db import connect
from .auth import require_auth
from .rbac import require_roles


router = APIRouter(prefix="/stats", tags=["stats"])


# =========================
# Auth helpers
# =========================
def _auth_employee(request: Request):
    user = require_auth(request)
    require_roles(user, ["employee", "supervisor", "admin"])  # الموظف والمشرف والأدمن
    return user


def _auth_admin(request: Request):
    user = require_auth(request)
    require_roles(user, ["admin"])  # الأدمن فقط
    return user


def _auth_stats(request: Request):
    user = require_auth(request)
    require_roles(user, ["admin", "supervisor", "employee"])
    return user


# =========================
# DB helpers
# =========================
def _q_scalar(sql: str, params: tuple = ()) -> int | float | None:
    # ✅ #1: try/finally لضمان إغلاق الـ connection حتى عند الخطأ
    con = connect()
    try:
        cur = con.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        if not row:
            return None
        return next(iter(row.values()))
    finally:
        con.close()


def _q_rows(sql: str, params: tuple = ()) -> list[dict]:
    # ✅ #1: try/finally لضمان إغلاق الـ connection حتى عند الخطأ
    con = connect()
    try:
        cur = con.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        return [dict(r) for r in rows]
    finally:
        con.close()


# =========================
# Excel export helpers (Enhanced)
# =========================
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="4F46E5")  # Indigo
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
CELL_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _ws_rtl(ws):
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass


def _style_header_row(ws, row: int = 1):
    ws.row_dimensions[row].height = 22
    for cell in ws[row]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER


def _style_body(ws, start_row: int = 2):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.data_type == "n":
                cell.alignment = CELL_ALIGN_CENTER
            else:
                cell.alignment = CELL_ALIGN


def _freeze_and_filter(ws, freeze_cell: str = "A2"):
    ws.freeze_panes = freeze_cell
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def _set_col_formats(ws, col_formats: dict):
    """
    col_formats: { "B": "0", "E": "yyyy-mm-dd hh:mm", "C": "0.00%" }
    Applies from row 2..max_row
    """
    for col_letter, fmt in col_formats.items():
        for r in range(2, ws.max_row + 1):
            ws[f"{col_letter}{r}"].number_format = fmt


def _wb_autofit(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def _safe_int(x: Any, default: int = 0) -> int:
    try:
        return int(x)
    except Exception:
        return default


def _safe_float(x: Any, default: float = 0.0) -> float:
    try:
        return float(x)
    except Exception:
        return default


def _safe_dt(x: Any) -> Any:
    """Remove timezone info from datetime objects so openpyxl can write them."""
    from datetime import datetime as _dt
    if isinstance(x, _dt) and x.tzinfo is not None:
        return x.replace(tzinfo=None)
    return x


# =========================
# Internal data builders
# =========================
def _overview_data() -> dict:
    total_conversations = _q_scalar("SELECT COUNT(*) FROM conversation") or 0
    total_messages = _q_scalar("SELECT COUNT(*) FROM message WHERE message_type='assistant'") or 0

    avg_msgs_per_conv = 0
    if total_conversations:
        avg_msgs_per_conv = round(total_messages / total_conversations, 2)

    answer_found_rate = _q_scalar(
        "SELECT ROUND(AVG(CASE WHEN answer_found IS NULL THEN 0 ELSE answer_found END)::numeric, 4) FROM message WHERE message_type='assistant'"
    )
    if answer_found_rate is None:
        answer_found_rate = 0.0

    avg_intent_conf = _q_scalar(
        # ✅ #2: فقط رسائل assistant — رسائل user ليس لها intent_conf
        "SELECT ROUND(AVG(CASE WHEN intent_conf IS NULL THEN 0 ELSE intent_conf END)::numeric, 4) FROM message WHERE message_type='assistant'"
    )
    if avg_intent_conf is None:
        avg_intent_conf = 0.0

    return {
        "total_conversations": int(total_conversations),
        "total_messages": int(total_messages),
        "avg_messages_per_conversation": avg_msgs_per_conv,
        "answer_found_rate": float(answer_found_rate),
        "avg_intent_conf": float(avg_intent_conf),
    }


def _daily_data(limit: int = 30) -> dict:
    limit = max(1, min(limit, 365))

    conv = _q_rows(
        """
        SELECT DATE(started_at) AS day, COUNT(*) AS conversations
        FROM conversation
        GROUP BY DATE(started_at)
        ORDER BY day DESC
        LIMIT %s
        """,
        (limit,),
    )

    msg = _q_rows(
        """
        SELECT DATE(created_at) AS day, COUNT(*) AS messages
        FROM message
        WHERE message_type = 'assistant'
        GROUP BY DATE(created_at)
        ORDER BY day DESC
        LIMIT %s
        """,
        (limit,),
    )

    return {
        "conversations_daily": list(reversed(conv)),
        "messages_daily": list(reversed(msg)),
    }


def _peak_hours_data(days: int = 30) -> dict:
    days = max(1, min(days, 365))

    rows = _q_rows(
        """
        SELECT LPAD(EXTRACT(HOUR FROM created_at)::text, 2, '0') AS hour, COUNT(*) AS total
        FROM message
        WHERE created_at >= NOW() + %s::interval
        GROUP BY hour
        ORDER BY hour
        """,
        (f"-{days} days",),
    )

    filled = {f"{h:02d}": 0 for h in range(24)}
    for r in rows:
        filled[str(r["hour"])] = int(r["total"])

    out = [{"hour": h, "total": filled[h]} for h in sorted(filled.keys())]
    return {"days": days, "hours": out}


def _top_intents_data(limit: int = 10, days: int = 30) -> dict:
    limit = max(1, min(limit, 50))
    days = max(1, min(days, 365))

    rows = _q_rows(
        """
        SELECT COALESCE(intent_pred, 'unknown') AS intent, COUNT(*) AS total
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + %s::interval
        GROUP BY intent
        ORDER BY total DESC
        LIMIT %s
        """,
        (f"-{days} days", limit),
    )
    return {"days": days, "top_intents": rows}


def _response_modes_data(days: int = 30) -> dict:
    days = max(1, min(days, 365))

    rows = _q_rows(
        """
        SELECT COALESCE(response_mode, 'unknown') AS mode, COUNT(*) AS total
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + %s::interval
        GROUP BY mode
        ORDER BY total DESC
        """,
        (f"-{days} days",),
    )
    return {"days": days, "response_modes": rows}


def _kb_usage_data(limit: int = 10, days: int = 30) -> dict:
    limit = max(1, min(limit, 50))
    days = max(1, min(days, 365))

    rows = _q_rows(
        """
        SELECT COALESCE(source_file, 'none') AS source_file, COUNT(*) AS total
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + %s::interval
        GROUP BY source_file
        ORDER BY total DESC
        LIMIT %s
        """,
        (f"-{days} days", limit),
    )
    return {"days": days, "kb_usage": rows}


def _quality_data(days: int = 30) -> dict:
    days = max(1, min(days, 365))

    answer_found_rate = _q_scalar(
        """
        SELECT ROUND(AVG(CASE WHEN answer_found IS NULL THEN 0 ELSE answer_found END)::numeric, 4)
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + %s::interval
        """,
        (f"-{days} days",),
    ) or 0.0

    avg_best_score = _q_scalar(
        """
        SELECT ROUND((AVG(CASE WHEN best_score IS NULL THEN 0 ELSE best_score END))::numeric, 2)
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + %s::interval
        """,
        (f"-{days} days",),
    ) or 0.0

    score_buckets = _q_rows(
        """
        SELECT
          CASE
            WHEN best_score IS NULL THEN 'null'
            WHEN best_score < 10 THEN '<10'
            WHEN best_score < 16 THEN '10-15'
            WHEN best_score < 25 THEN '16-24'
            ELSE '25+'
          END AS bucket,
          COUNT(*) AS total
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + %s::interval
        GROUP BY bucket
        ORDER BY total DESC
        """,
        (f"-{days} days",),
    )

    return {
        "days": days,
        "answer_found_rate": float(answer_found_rate),
        "avg_best_score": float(avg_best_score),
        "best_score_buckets": score_buckets,
    }


def _complaints_summary_data(days: int = 30) -> dict:
    days = max(1, min(days, 365))

    rows = _q_rows(
        """
        SELECT
          COALESCE(intent_pred, 'unknown') AS category,
          COUNT(*) AS total
        FROM message
        WHERE created_at >= NOW() + %s::interval
          AND (
            LOWER(COALESCE(intent_pred,'')) LIKE 'complaint%%'
            OR LOWER(COALESCE(response_mode,'')) LIKE '%%complaint%%'
          )
        GROUP BY intent_pred
        ORDER BY total DESC
        """,
        (f"-{days} days",),
    )

    total = sum(int(r["total"]) for r in rows) if rows else 0

    breakdown = []
    for r in rows:
        t = int(r["total"])
        pct = (t / total * 100.0) if total else 0.0
        breakdown.append({"category": r["category"], "total": t, "percent": round(pct, 2)})

    top = breakdown[0]["category"] if breakdown else "—"

    return {
        "days": days,
        "total_complaints": total,
        "top_complaint": top,
        "breakdown": breakdown,
    }


def _employees_activity_data(days: int = 30) -> dict:
    days = max(1, min(int(days or 30), 365))

    rows = _q_rows(
        """
        SELECT
            u.user_id,
            u.full_name,
            u.role,
            COUNT(s.session_id) AS logins,
            MAX(s.last_activity) AS last_activity,
            SUM(
                CASE
                    WHEN s.last_activity IS NOT NULL
                    THEN EXTRACT(EPOCH FROM (s.last_activity - s.created_at)) / 60
                    ELSE 0
                END
            ) AS active_minutes
        FROM app_user u
        LEFT JOIN user_session s
            ON s.user_id = u.user_id
            AND s.created_at >= NOW() + %s::interval
        WHERE u.role IN ('employee','supervisor','admin')
        GROUP BY u.user_id
        ORDER BY logins DESC
        """,
        (f"-{days} days",),
    )

    out = []
    for r in rows:
        minutes = int(r.get("active_minutes") or 0)

        if minutes > 120:
            status = "نشط"
        elif minutes > 10:
            status = "قليل النشاط"
        else:
            status = "غير مستخدم"

        out.append(
            {
                "user_id": int(r["user_id"]),
                "name": r.get("full_name") or "",
                "role": r.get("role") or "",  # اختياري (لو بدك تعرضه بالفرونت)
                "logins": int(r.get("logins") or 0),
                "last_activity": r.get("last_activity"),  # may be None
                "active_minutes": minutes,
                "status": status,
            }
        )

    # ترتيب حسب النشاط (desc)
    out.sort(key=lambda x: int(x.get("active_minutes") or 0), reverse=True)

    return {"days": days, "employees": out}


# =========================
# Endpoints
# =========================
@router.get("/overview")
def overview(_: dict = Depends(_auth_stats)):
    return _overview_data()


@router.get("/daily")
def daily(limit: int = 30, _: dict = Depends(_auth_stats)):
    return _daily_data(limit=limit)


@router.get("/peak-hours")
def peak_hours(days: int = 30, _: dict = Depends(_auth_stats)):
    return _peak_hours_data(days=days)


@router.get("/top-intents")
def top_intents(limit: int = 10, days: int = 30, _: dict = Depends(_auth_stats)):
    return _top_intents_data(limit=limit, days=days)


@router.get("/response-modes")
def response_modes(days: int = 30, _: dict = Depends(_auth_stats)):
    return _response_modes_data(days=days)


@router.get("/kb-usage")
def kb_usage(limit: int = 10, days: int = 30, _: dict = Depends(_auth_stats)):
    return _kb_usage_data(limit=limit, days=days)


@router.get("/quality")
def quality(days: int = 30, _: dict = Depends(_auth_stats)):
    return _quality_data(days=days)


@router.get("/complaints-summary")
def complaints_summary(days: int = 30, _: dict = Depends(_auth_stats)):
    return _complaints_summary_data(days=days)


@router.get("/employees-activity")
def employees_activity(days: int = 30, _: dict = Depends(_auth_stats)):
    return _employees_activity_data(days=days)


@router.get("/low-rated-conversations")
def low_rated_conversations(days: int = 30, threshold: int = 2, limit: int = 50, _: dict = Depends(_auth_stats)):
    days = max(1, min(int(days or 30), 365))
    threshold = max(1, min(int(threshold or 2), 5))
    limit = max(1, min(int(limit or 50), 200))

    convs = _q_rows(
        """
        SELECT
          conversation_id,
          ROUND(AVG(rating::numeric), 2) AS avg_stars,
          COUNT(*) AS ratings_count,
          MAX(submitted_at) AS last_rated_at
        FROM feedback
        WHERE submitted_at >= NOW() + %s::interval
          AND feedback_type='stars'
          AND conversation_id IS NOT NULL
        GROUP BY conversation_id
        HAVING ROUND(AVG(rating::numeric), 2) <= %s
        ORDER BY avg_stars ASC, last_rated_at DESC
        LIMIT %s
        """,
        (f"-{days} days", threshold, limit),
    )

    out = []
    for c in convs:
        cid = int(c["conversation_id"])
        preview = _q_rows(
            """
            SELECT message_type, message_text, response_text, created_at
            FROM message
            WHERE conversation_id=%s
            ORDER BY message_id DESC
            LIMIT 8
            """,
            (cid,),
        )
        out.append({**c, "preview": list(reversed(preview))})

    return {"days": days, "threshold": threshold, "items": out}


# =========================
# Stars analytics
# =========================
@router.get("/stars-overview")
def stars_overview(days: int = 30, _: dict = Depends(_auth_stats)):
    days = max(1, min(int(days or 30), 365))

    total_ratings = _q_scalar(
        """
        SELECT COUNT(*)
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          AND submitted_at >= NOW() + %s::interval
        """,
        (f"-{days} days",),
    ) or 0

    avg_stars = _q_scalar(
        """
        SELECT ROUND(AVG(rating::numeric), 2)
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          AND submitted_at >= NOW() + %s::interval
        """,
        (f"-{days} days",),
    )
    if avg_stars is None:
        avg_stars = 0.0

    dist = _q_rows(
        """
        SELECT CAST(rating AS INT) AS stars, COUNT(*) AS total
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          AND submitted_at >= NOW() + %s::interval
        GROUP BY CAST(rating AS INT)
        ORDER BY stars
        """,
        (f"-{days} days",),
    )

    dist_map = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for r in dist:
        s = int(r["stars"] or 0)
        if s in dist_map:
            dist_map[s] = int(r["total"] or 0)

    satisfaction_rate = 0.0
    if total_ratings:
        high = dist_map[4] + dist_map[5]
        satisfaction_rate = round((high / total_ratings) * 100.0, 1)

    return {
        "days": days,
        "total_ratings": int(total_ratings),
        "avg_stars": float(avg_stars),
        "satisfaction_rate": float(satisfaction_rate),
        "distribution": [
            {"stars": 1, "total": dist_map[1]},
            {"stars": 2, "total": dist_map[2]},
            {"stars": 3, "total": dist_map[3]},
            {"stars": 4, "total": dist_map[4]},
            {"stars": 5, "total": dist_map[5]},
        ],
    }


@router.get("/stars-weekly")
def stars_weekly(days: int = 30, _: dict = Depends(_auth_stats)):
    days = max(1, min(int(days or 30), 365))

    rows = _q_rows(
        """
        SELECT
          DATE(submitted_at) AS day,
          ROUND(AVG(rating::numeric), 2) AS avg_stars,
          COUNT(*) AS count
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          AND submitted_at >= NOW() + %s::interval
        GROUP BY DATE(submitted_at)
        ORDER BY day ASC
        """,
        (f"-{days} days",),
    )

    return {"days": days, "by_day": rows}


@router.get("/recent-feedback")
def recent_feedback(days: int = 30, limit: int = 10, _: dict = Depends(_auth_stats)):
    days = max(1, min(int(days or 30), 365))
    limit = max(1, min(int(limit or 10), 50))

    rows = _q_rows(
        """
        SELECT
          conversation_id,
          CAST(rating AS INT) AS stars,
          COALESCE(comments, '') AS comment,
          submitted_at
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          AND submitted_at >= NOW() + %s::interval
        ORDER BY submitted_at DESC
        LIMIT %s
        """,
        (f"-{days} days", limit),
    )

    items = [
        {
            "conversation_id": int(r["conversation_id"]),
            "stars": int(r["stars"] or 0),
            "comment": r.get("comment") or None,
            "submitted_at": r["submitted_at"],
        }
        for r in rows
    ]

    return {"days": days, "items": items}


# =========================
# Range helpers + endpoints
# =========================
def _date_range_where(from_date: str | None, to_date: str | None) -> tuple[str, tuple]:
    where = []
    params: list[str] = []

    if from_date:
        where.append("submitted_at >= %s::timestamp")
        params.append(f"{from_date} 00:00:00")
    if to_date:
        where.append("submitted_at <= %s::timestamp")
        params.append(f"{to_date} 23:59:59")

    clause = " AND " + " AND ".join(where) if where else ""
    return clause, tuple(params)


@router.get("/conversation-ratings-summary-range")
def conversation_ratings_summary_range(
    from_date: str | None = Query(None, description="YYYY-MM-DD"),
    to_date: str | None = Query(None, description="YYYY-MM-DD"),
    _: dict = Depends(_auth_stats),
):
    clause, params = _date_range_where(from_date, to_date)

    total_ratings = _q_scalar(
        f"""
        SELECT COUNT(*)
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          {clause}
        """,
        params,
    ) or 0

    avg_rating = _q_scalar(
        f"""
        SELECT ROUND(AVG(rating::numeric), 2)
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          {clause}
        """,
        params,
    )
    if avg_rating is None:
        avg_rating = 0.0

    dist = _q_rows(
        f"""
        SELECT CAST(rating AS INT) AS stars, COUNT(*) AS total
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          {clause}
        GROUP BY CAST(rating AS INT)
        ORDER BY stars
        """,
        params,
    )

    dist_map = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for r in dist:
        s = int(r["stars"] or 0)
        if s in dist_map:
            dist_map[s] = int(r["total"] or 0)

    satisfaction_rate = 0.0
    if total_ratings:
        high = dist_map[4] + dist_map[5]
        satisfaction_rate = round((high / total_ratings) * 100, 1)

    return {
        "from": from_date,
        "to": to_date,
        "total_ratings": int(total_ratings),
        "avg_rating": float(avg_rating),
        "satisfaction_rate": float(satisfaction_rate),
        "star_distribution": [
            {"stars": 1, "total": dist_map[1]},
            {"stars": 2, "total": dist_map[2]},
            {"stars": 3, "total": dist_map[3]},
            {"stars": 4, "total": dist_map[4]},
            {"stars": 5, "total": dist_map[5]},
        ],
    }


@router.get("/conversation-ratings-daily-range")
def conversation_ratings_daily_range(
    from_date: str | None = Query(None, description="YYYY-MM-DD"),
    to_date: str | None = Query(None, description="YYYY-MM-DD"),
    _: dict = Depends(_auth_stats),
):
    clause, params = _date_range_where(from_date, to_date)

    rows = _q_rows(
        f"""
        SELECT
          DATE(submitted_at) AS day,
          ROUND(AVG(rating::numeric), 2) AS avg_rating,
          COUNT(*) AS total
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          {clause}
        GROUP BY DATE(submitted_at)
        ORDER BY day ASC
        """,
        params,
    )

    return {"from": from_date, "to": to_date, "daily": rows}


@router.get("/recent-ratings-range")
def recent_ratings_range(
    limit: int = 50,
    from_date: str | None = Query(None, description="YYYY-MM-DD"),
    to_date: str | None = Query(None, description="YYYY-MM-DD"),
    _: dict = Depends(_auth_stats),
):
    limit = max(1, min(int(limit or 50), 200))
    clause, params = _date_range_where(from_date, to_date)

    rows = _q_rows(
        f"""
        SELECT
          conversation_id,
          CAST(rating AS INT) AS stars,
          COALESCE(comments, '') AS comment,
          submitted_at
        FROM feedback
        WHERE feedback_type='stars'
          AND conversation_id IS NOT NULL
          {clause}
        ORDER BY submitted_at DESC
        LIMIT %s
        """,
        params + (limit,),
    )

    return {"from": from_date, "to": to_date, "items": rows}


# =========================
# ✅ EXPORT: Monthly report (Excel/PDF)
# =========================
# =========================
# Export helper wrappers
# =========================
def _stars_overview_data(days: int = 30) -> dict:
    days = max(1, min(days, 365))
    total_ratings = _q_scalar(
        """SELECT COUNT(*) FROM feedback
           WHERE feedback_type='stars' AND conversation_id IS NOT NULL
             AND submitted_at >= NOW() + %s::interval""", (f"-{days} days",)) or 0
    avg_stars = _q_scalar(
        """SELECT ROUND(AVG(rating::numeric), 2) FROM feedback
           WHERE feedback_type='stars' AND conversation_id IS NOT NULL
             AND submitted_at >= NOW() + %s::interval""", (f"-{days} days",)) or 0.0
    dist = _q_rows(
        """SELECT CAST(rating AS INT) AS stars, COUNT(*) AS total FROM feedback
           WHERE feedback_type='stars' AND conversation_id IS NOT NULL
             AND submitted_at >= NOW() + %s::interval
           GROUP BY CAST(rating AS INT) ORDER BY stars""", (f"-{days} days",))
    dist_map = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for r in dist:
        s = int(r["stars"] or 0)
        if s in dist_map:
            dist_map[s] = int(r["total"] or 0)
    sat = 0.0
    if total_ratings:
        sat = round((dist_map[4] + dist_map[5]) / total_ratings * 100, 1)
    return {"days": days, "total_ratings": int(total_ratings), "avg_stars": float(avg_stars),
            "satisfaction_rate": sat,
            "distribution": [{"stars": s, "total": dist_map[s]} for s in range(1, 6)]}


def _stars_daily_data(days: int = 30) -> dict:
    days = max(1, min(days, 365))
    rows = _q_rows(
        """SELECT DATE(submitted_at) AS day,
                  ROUND(AVG(rating::numeric), 2) AS avg_stars,
                  COUNT(*) AS count
           FROM feedback
           WHERE feedback_type='stars' AND conversation_id IS NOT NULL
             AND submitted_at >= NOW() + %s::interval
           GROUP BY DATE(submitted_at) ORDER BY day ASC""", (f"-{days} days",))
    return {"days": days, "by_day": rows}


def _stars_recent_data(limit: int = 100, days: int = 30) -> dict:
    limit = max(1, min(limit, 500))
    days  = max(1, min(days, 365))
    rows = _q_rows(
        """SELECT conversation_id, CAST(rating AS INT) AS stars,
                  COALESCE(comments, '') AS comment, submitted_at
           FROM feedback
           WHERE feedback_type='stars' AND conversation_id IS NOT NULL
             AND submitted_at >= NOW() + %s::interval
           ORDER BY submitted_at DESC LIMIT %s""", (f"-{days} days", limit))
    items = [{"conversation_id": int(r["conversation_id"]), "stars": int(r["stars"] or 0),
              "comment": r.get("comment") or None, "submitted_at": r["submitted_at"]} for r in rows]
    return {"days": days, "items": items}


def _low_rated_data(days: int = 30, threshold: int = 2, limit: int = 100) -> dict:
    days      = max(1, min(days, 365))
    threshold = max(1, min(threshold, 5))
    limit     = max(1, min(limit, 500))
    convs = _q_rows(
        """SELECT conversation_id, ROUND(AVG(rating::numeric), 2) AS avg_stars,
                  COUNT(*) AS ratings_count, MAX(submitted_at) AS last_rated_at
           FROM feedback
           WHERE submitted_at >= NOW() + %s::interval AND feedback_type='stars'
             AND conversation_id IS NOT NULL
           GROUP BY conversation_id HAVING ROUND(AVG(rating::numeric), 2) <= %s
           ORDER BY avg_stars ASC, last_rated_at DESC LIMIT %s""",
        (f"-{days} days", threshold, limit))
    out = []
    for c in convs:
        cid = int(c["conversation_id"])
        preview = _q_rows(
            """SELECT message_type, message_text, response_text, created_at
               FROM message WHERE conversation_id=%s ORDER BY message_id DESC LIMIT 8""", (cid,))
        out.append({**c, "preview": list(reversed(preview))})
    return {"days": days, "threshold": threshold, "items": out}


@router.get("/export/monthly.xlsx")
def export_monthly_excel(days: int = 30, _: dict = Depends(_auth_stats)):
    days = max(1, min(int(days or 30), 365))

    ov            = _overview_data()
    daily_data    = _daily_data(limit=min(days, 365))
    intents       = _top_intents_data(limit=20, days=days)
    peak          = _peak_hours_data(days=days)
    modes         = _response_modes_data(days=days)
    kb            = _kb_usage_data(limit=20, days=days)
    qual          = _quality_data(days=days)
    complaints    = _complaints_summary_data(days=days)
    stars_ov      = _stars_overview_data(days=days)
    stars_daily   = _stars_daily_data(days=days)
    stars_recent  = _stars_recent_data(limit=100, days=days)
    low_rated     = _low_rated_data(days=days, threshold=2, limit=100)
    emp           = _employees_activity_data(days=days)

    wb = Workbook()

    # ── Sheet 1: ملخص عام ─────────────────────────────────────
    ws = wb.active
    ws.title = "ملخص عام"
    _ws_rtl(ws)
    ws.append(["البند", "القيمة"])
    for k, v in [
        ("إجمالي المحادثات",              ov.get("total_conversations")),
        ("إجمالي الرسائل",                ov.get("total_messages")),
        ("متوسط الرسائل / محادثة",        ov.get("avg_messages_per_conversation")),
        ("نسبة العثور على إجابة",         ov.get("answer_found_rate")),
        ("متوسط ثقة التصنيف",             ov.get("avg_intent_conf")),
        ("",                               ""),
        ("إجمالي التقييمات (نجوم)",       stars_ov.get("total_ratings", 0)),
        ("متوسط النجوم",                   stars_ov.get("avg_stars", 0)),
        ("نسبة الرضا (4-5 نجوم)",         stars_ov.get("satisfaction_rate", 0) / 100.0),
        ("",                               ""),
        ("إجمالي الشكاوى",                complaints.get("total_complaints", 0)),
        ("أكثر شكوى",                     complaints.get("top_complaint", "—")),
        ("",                               ""),
        ("الفترة (أيام)",                 days),
        ("تاريخ التوليد (UTC)",            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")),
    ]:
        ws.append([k, v])

    _style_header_row(ws, 1)
    _style_body(ws, 2)
    _freeze_and_filter(ws, "A2")
    for r in range(2, ws.max_row + 1):
        lbl = str(ws[f"A{r}"].value or "")
        if "نسبة" in lbl or "ثقة" in lbl:
            ws[f"B{r}"].number_format = "0.00%"
        if "متوسط النجوم" in lbl:
            ws[f"B{r}"].number_format = "0.00"
    _wb_autofit(ws)

    # ── Sheet 2: يومي ─────────────────────────────────────────
    ws2 = wb.create_sheet("يومي")
    _ws_rtl(ws2)
    ws2.append(["اليوم", "محادثات", "رسائل"])
    conv_map = {r["day"]: _safe_int(r.get("conversations")) for r in (daily_data.get("conversations_daily") or [])}
    msg_map  = {r["day"]: _safe_int(r.get("messages"))      for r in (daily_data.get("messages_daily")      or [])}
    for d in sorted(set(list(conv_map) + list(msg_map))):
        ws2.append([d, conv_map.get(d, 0), msg_map.get(d, 0)])
    _style_header_row(ws2, 1); _style_body(ws2, 2)
    _freeze_and_filter(ws2, "A2"); _set_col_formats(ws2, {"B": "0", "C": "0"})
    _wb_autofit(ws2)

    # ── Sheet 3: أوقات الذروة ─────────────────────────────────
    ws3 = wb.create_sheet("أوقات الذروة")
    _ws_rtl(ws3)
    ws3.append(["الساعة", "عدد الرسائل"])
    for r in (peak.get("hours") or []):
        ws3.append([f"{r.get('hour')}:00", _safe_int(r.get("total"))])
    _style_header_row(ws3, 1); _style_body(ws3, 2)
    _freeze_and_filter(ws3, "A2"); _set_col_formats(ws3, {"B": "0"})
    _wb_autofit(ws3)

    # ── Sheet 4: التصنيفات (Intents) ──────────────────────────
    ws4 = wb.create_sheet("التصنيفات")
    _ws_rtl(ws4)
    ws4.append(["التصنيف", "العدد"])
    total_intents = sum(_safe_int(r.get("total")) for r in (intents.get("top_intents") or []))
    for r in (intents.get("top_intents") or []):
        cnt = _safe_int(r.get("total"))
        pct = round(cnt / total_intents * 100, 1) if total_intents else 0
        ws4.append([r.get("intent"), cnt, f"{pct}%"])
    if ws4.max_row > 1:
        ws4["C1"] = "النسبة"
    _style_header_row(ws4, 1); _style_body(ws4, 2)
    _freeze_and_filter(ws4, "A2"); _set_col_formats(ws4, {"B": "0"})
    _wb_autofit(ws4)

    # ── Sheet 5: أوضاع الردود ─────────────────────────────────
    ws5 = wb.create_sheet("أوضاع الردود")
    _ws_rtl(ws5)
    ws5.append(["وضع الرد", "العدد", "النسبة"])
    total_modes = sum(_safe_int(r.get("total")) for r in (modes.get("response_modes") or []))
    for r in (modes.get("response_modes") or []):
        cnt = _safe_int(r.get("total"))
        pct = round(cnt / total_modes * 100, 1) if total_modes else 0
        ws5.append([r.get("mode"), cnt, pct / 100.0])
    _style_header_row(ws5, 1); _style_body(ws5, 2)
    _freeze_and_filter(ws5, "A2")
    _set_col_formats(ws5, {"B": "0", "C": "0.00%"})
    _wb_autofit(ws5)

    # ── Sheet 6: قاعدة المعرفة ────────────────────────────────
    ws6 = wb.create_sheet("قاعدة المعرفة")
    _ws_rtl(ws6)
    ws6.append(["ملف المصدر", "عدد الاستخدامات"])
    for r in (kb.get("kb_usage") or []):
        ws6.append([r.get("source_file"), _safe_int(r.get("total"))])
    _style_header_row(ws6, 1); _style_body(ws6, 2)
    _freeze_and_filter(ws6, "A2"); _set_col_formats(ws6, {"B": "0"})
    _wb_autofit(ws6)

    # ── Sheet 7: جودة الردود ──────────────────────────────────
    ws7 = wb.create_sheet("جودة الردود")
    _ws_rtl(ws7)
    ws7.append(["المؤشر", "القيمة"])
    ws7.append(["نسبة العثور على إجابة", _safe_float(qual.get("answer_found_rate"))])
    ws7.append(["متوسط أفضل تطابق (Score)", _safe_float(qual.get("avg_best_score"))])
    ws7.append(["", ""])
    ws7.append(["توزيع درجات التطابق", ""])
    ws7.append(["الفئة", "عدد الرسائل"])
    for r in (qual.get("best_score_buckets") or []):
        ws7.append([r.get("bucket"), _safe_int(r.get("total"))])
    _style_header_row(ws7, 1); _style_body(ws7, 2)
    _freeze_and_filter(ws7, "A2")
    ws7["B2"].number_format = "0.00%"
    ws7["B3"].number_format = "0.00"
    _wb_autofit(ws7)

    # ── Sheet 8: الشكاوى ──────────────────────────────────────
    ws8 = wb.create_sheet("الشكاوى")
    _ws_rtl(ws8)
    ws8.append(["البند", "القيمة"])
    ws8.append(["إجمالي الشكاوى", _safe_int(complaints.get("total_complaints"))])
    ws8.append(["أكثر شكوى", complaints.get("top_complaint")])
    ws8.append(["", ""])
    ws8.append(["التصنيف", "العدد", "النسبة"])
    for r in (complaints.get("breakdown") or []):
        ws8.append([r.get("category"), _safe_int(r.get("total")), _safe_float(r.get("percent")) / 100.0])
    _style_header_row(ws8, 1); _style_header_row(ws8, 5)
    _style_body(ws8, 2); _freeze_and_filter(ws8, "A2")
    _set_col_formats(ws8, {"B": "0", "C": "0.00%"})
    _wb_autofit(ws8)

    # ── Sheet 9: تقييمات النجوم ───────────────────────────────
    ws9 = wb.create_sheet("تقييمات النجوم")
    _ws_rtl(ws9)
    ws9.append(["البند", "القيمة"])
    ws9.append(["إجمالي التقييمات",  _safe_int(stars_ov.get("total_ratings"))])
    ws9.append(["متوسط النجوم",       _safe_float(stars_ov.get("avg_stars"))])
    ws9.append(["نسبة الرضا (4-5)",   _safe_float(stars_ov.get("satisfaction_rate")) / 100.0])
    ws9.append(["", ""])
    ws9.append(["توزيع النجوم", ""])
    ws9.append(["النجوم", "العدد"])
    for r in (stars_ov.get("distribution") or []):
        ws9.append([f"{'★' * int(r.get('stars', 0))} ({r.get('stars')})", _safe_int(r.get("total"))])
    _style_header_row(ws9, 1); _style_header_row(ws9, 6)
    _style_body(ws9, 2); _freeze_and_filter(ws9, "A2")
    ws9["B2"].number_format = "0"
    ws9["B3"].number_format = "0.00"
    ws9["B4"].number_format = "0.00%"
    _wb_autofit(ws9)

    # ── Sheet 10: تقييمات يومية ───────────────────────────────
    ws10 = wb.create_sheet("تقييمات يومية")
    _ws_rtl(ws10)
    ws10.append(["اليوم", "متوسط النجوم", "عدد التقييمات"])
    for r in (stars_daily.get("by_day") or []):
        ws10.append([r.get("day"), _safe_float(r.get("avg_stars")), _safe_int(r.get("count"))])
    _style_header_row(ws10, 1); _style_body(ws10, 2)
    _freeze_and_filter(ws10, "A2")
    _set_col_formats(ws10, {"B": "0.00", "C": "0"})
    _wb_autofit(ws10)

    # ── Sheet 11: آخر التعليقات ───────────────────────────────
    ws11 = wb.create_sheet("آخر التعليقات")
    _ws_rtl(ws11)
    ws11.append(["رقم المحادثة", "النجوم", "التعليق", "التاريخ"])
    for r in (stars_recent.get("items") or []):
        ws11.append([
            _safe_int(r.get("conversation_id")),
            _safe_int(r.get("stars")),
            r.get("comment") or "",
            _safe_dt(r.get("submitted_at")) or "",
        ])
    _style_header_row(ws11, 1); _style_body(ws11, 2)
    _freeze_and_filter(ws11, "A2"); _set_col_formats(ws11, {"A": "0", "B": "0"})
    _wb_autofit(ws11)

    # ── Sheet 12: محادثات بتقييم منخفض ───────────────────────
    ws12 = wb.create_sheet("تقييم منخفض")
    _ws_rtl(ws12)
    ws12.append(["رقم المحادثة", "متوسط النجوم", "عدد التقييمات", "آخر تقييم", "معاينة السؤال", "معاينة الرد"])
    for c in (low_rated.get("items") or []):
        preview = (c.get("preview") or [])
        last_q = next((m.get("message_text") or "" for m in reversed(preview) if m.get("message_type") == "user"), "")
        last_a = next((m.get("response_text") or "" for m in reversed(preview) if m.get("message_type") == "assistant"), "")
        ws12.append([
            _safe_int(c.get("conversation_id")),
            _safe_float(c.get("avg_stars")),
            _safe_int(c.get("ratings_count")),
            _safe_dt(c.get("last_rated_at")) or "",
            (last_q or "")[:200],
            (last_a or "")[:300],
        ])
    _style_header_row(ws12, 1); _style_body(ws12, 2)
    _freeze_and_filter(ws12, "A2"); _set_col_formats(ws12, {"A": "0", "B": "0.00", "C": "0"})
    _wb_autofit(ws12)

    # ── Sheet 13: نشاط الموظفين ───────────────────────────────
    ws13 = wb.create_sheet("نشاط الموظفين")
    _ws_rtl(ws13)
    emps = sorted(emp.get("employees", []), key=lambda x: int(x.get("active_minutes") or 0), reverse=True)
    ws13.append(["#", "رقم الموظف", "الاسم", "الدور", "عدد الدخول", "آخر نشاط", "دقائق النشاط", "الحالة"])
    for idx, e in enumerate(emps, 1):
        ws13.append([idx, e.get("user_id"), e.get("name"), e.get("role"),
                     _safe_int(e.get("logins")), _safe_dt(e.get("last_activity")), _safe_int(e.get("active_minutes")), e.get("status")])
    _style_header_row(ws13, 1); _style_body(ws13, 2)
    _freeze_and_filter(ws13, "A2")
    _set_col_formats(ws13, {"A": "0", "B": "0", "E": "0", "G": "0"})
    for r in range(2, ws13.max_row + 1):
        ws13[f"F{r}"].number_format = "yyyy-mm-dd hh:mm"
    status_range = f"H2:H{ws13.max_row}"
    ws13.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"نشط"'], fill=PatternFill("solid", fgColor="DCFCE7")))
    ws13.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"قليل النشاط"'], fill=PatternFill("solid", fgColor="FEF9C3")))
    ws13.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"غير مستخدم"'], fill=PatternFill("solid", fgColor="F3F4F6")))
    _wb_autofit(ws13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"full_report_{days}d_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/monthly.pdf")
def export_monthly_pdf(days: int = 30, _: dict = Depends(_auth_stats)):
    days = max(1, min(int(days or 30), 365))

    ov           = _overview_data()
    daily_data   = _daily_data(limit=min(days, 30))
    intents      = _top_intents_data(limit=15, days=days)
    peak         = _peak_hours_data(days=days)
    modes        = _response_modes_data(days=days)
    kb           = _kb_usage_data(limit=10, days=days)
    qual         = _quality_data(days=days)
    complaints   = _complaints_summary_data(days=days)
    stars_ov     = _stars_overview_data(days=days)
    stars_daily  = _stars_daily_data(days=days)
    stars_recent = _stars_recent_data(limit=10, days=days)
    low_rated    = _low_rated_data(days=days, threshold=2, limit=20)
    emp          = _employees_activity_data(days=days)

    buf = io.BytesIO()
    cv = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    # ── helpers ───────────────────────────────────────────────
    y_state = [H - 50]

    def new_page():
        cv.showPage()
        y_state[0] = H - 50

    def check_space(needed: float = 30):
        if y_state[0] < needed + 40:
            new_page()

    def draw_title(txt: str):
        check_space(40)
        cv.setFont(_pdf_font(bold=True), 13)
        cv.setFillColorRGB(0.24, 0.27, 0.83)
        cv.drawRightString(W - 40, y_state[0], _ar(txt))
        y_state[0] -= 4
        cv.setStrokeColorRGB(0.24, 0.27, 0.83)
        cv.setLineWidth(1)
        cv.line(40, y_state[0], W - 40, y_state[0])
        y_state[0] -= 14
        cv.setFillColorRGB(0, 0, 0)

    def draw_row(label: str, value: str, indent: int = 60):
        check_space(16)
        cv.setFont(_pdf_font(bold=False), 9)
        cv.drawRightString(W - indent, y_state[0], _ar(str(value)[:80]))
        cv.setFont(_pdf_font(bold=True), 9)
        cv.drawRightString(W - indent - 10, y_state[0], _ar(f"{label}:"))
        y_state[0] -= 13

    def draw_line(txt: str, indent: int = 60, bold: bool = False):
        check_space(14)
        cv.setFont(_pdf_font(bold=bold), 9)
        cv.drawRightString(W - indent, y_state[0], _ar(str(txt)[:100]))
        y_state[0] -= 13

    def draw_table_header(cols: list, widths: list, x0: int = 40):
        check_space(18)
        cv.setFillColorRGB(0.31, 0.28, 0.90)
        cv.rect(x0, y_state[0] - 2, sum(widths), 16, fill=1, stroke=0)
        cv.setFillColorRGB(1, 1, 1)
        cv.setFont(_pdf_font(bold=True), 8)
        x = x0 + sum(widths)
        for col, w2 in zip(cols, widths):
            x -= w2
            cv.drawCentredString(x + w2 / 2, y_state[0] + 1, _ar(str(col)[:24]))
        y_state[0] -= 16
        cv.setFillColorRGB(0, 0, 0)

    def draw_table_row(vals: list, widths: list, x0: int = 40, alt: bool = False):
        check_space(14)
        if alt:
            cv.setFillColorRGB(0.96, 0.96, 1.0)
            cv.rect(x0, y_state[0] - 2, sum(widths), 13, fill=1, stroke=0)
            cv.setFillColorRGB(0, 0, 0)
        cv.setFont(_pdf_font(), 8)
        x = x0 + sum(widths)
        for v, w2 in zip(vals, widths):
            x -= w2
            cv.drawCentredString(x + w2 / 2, y_state[0] + 1, _ar(str(v)[:30]))
        y_state[0] -= 13

    def spacer(n: float = 8):
        y_state[0] -= n

    # ── Cover ─────────────────────────────────────────────────
    cv.setFillColorRGB(0.24, 0.27, 0.83)
    cv.rect(0, H - 90, W, 90, fill=1, stroke=0)
    cv.setFillColorRGB(1, 1, 1)
    cv.setFont("Helvetica-Bold", 18)
    cv.drawCentredString(W / 2, H - 45, f"Full System Report  |  Last {days} Days")
    cv.setFont("Helvetica", 10)
    cv.drawCentredString(W / 2, H - 65, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    cv.setFillColorRGB(0, 0, 0)
    y_state[0] = H - 110

    # ── 1. ملخص عام ────────────────────────────────────────────
    draw_title("1. ملخص عام  (Overview)")
    draw_row("إجمالي المحادثات",       str(ov.get("total_conversations", 0)))
    draw_row("إجمالي الرسائل",         str(ov.get("total_messages", 0)))
    draw_row("متوسط الرسائل/محادثة",   str(ov.get("avg_messages_per_conversation", 0)))
    draw_row("نسبة العثور على إجابة",  f"{round(_safe_float(ov.get('answer_found_rate')) * 100, 1)}%")
    draw_row("متوسط ثقة التصنيف",      f"{round(_safe_float(ov.get('avg_intent_conf')) * 100, 1)}%")
    spacer()

    # ── 2. جودة الردود ─────────────────────────────────────────
    draw_title("2. جودة الردود  (Quality)")
    draw_row("نسبة العثور على إجابة",  f"{round(_safe_float(qual.get('answer_found_rate')) * 100, 1)}%")
    draw_row("متوسط أفضل تطابق",       f"{_safe_float(qual.get('avg_best_score')):.2f}")
    spacer(4)
    draw_table_header(["فئة التطابق", "عدد الرسائل"], [140, 100])
    for i, r in enumerate(qual.get("best_score_buckets") or []):
        draw_table_row([r.get("bucket", ""), _safe_int(r.get("total"))], [140, 100], alt=bool(i % 2))
    spacer()

    # ── 3. أوقات الذروة ────────────────────────────────────────
    draw_title("3. أوقات الذروة  (Peak Hours)")
    hours = peak.get("hours") or []
    if hours:
        top5 = sorted(hours, key=lambda x: _safe_int(x.get("total")), reverse=True)[:8]
        draw_table_header(["الساعة", "عدد الرسائل"], [120, 120])
        for i, r in enumerate(top5):
            draw_table_row([f"{r.get('hour')}:00", _safe_int(r.get("total"))], [120, 120], alt=bool(i % 2))
    spacer()

    # ── 4. التصنيفات ───────────────────────────────────────────
    draw_title("4. أكثر التصنيفات  (Top Intents)")
    top_list = (intents.get("top_intents") or [])[:15]
    total_i = sum(_safe_int(r.get("total")) for r in top_list)
    draw_table_header(["التصنيف", "العدد", "النسبة"], [230, 80, 80])
    for i, r in enumerate(top_list):
        cnt = _safe_int(r.get("total"))
        pct = f"{round(cnt/total_i*100,1)}%" if total_i else "0%"
        draw_table_row([r.get("intent","")[:35], cnt, pct], [230, 80, 80], alt=bool(i % 2))
    spacer()

    # ── 5. أوضاع الردود ────────────────────────────────────────
    draw_title("5. أوضاع الردود  (Response Modes)")
    total_m = sum(_safe_int(r.get("total")) for r in (modes.get("response_modes") or []))
    draw_table_header(["وضع الرد", "العدد", "النسبة"], [200, 80, 80])
    for i, r in enumerate(modes.get("response_modes") or []):
        cnt = _safe_int(r.get("total"))
        pct = f"{round(cnt/total_m*100,1)}%" if total_m else "0%"
        draw_table_row([r.get("mode","")[:30], cnt, pct], [200, 80, 80], alt=bool(i % 2))
    spacer()

    # ── 6. قاعدة المعرفة ───────────────────────────────────────
    draw_title("6. استخدام قاعدة المعرفة  (KB Usage)")
    draw_table_header(["الملف / المصدر", "عدد الاستخدامات"], [280, 120])
    for i, r in enumerate(kb.get("kb_usage") or []):
        draw_table_row([r.get("source_file","")[:40], _safe_int(r.get("total"))], [280, 120], alt=bool(i % 2))
    spacer()

    # ── 7. الشكاوى ─────────────────────────────────────────────
    draw_title("7. ملخص الشكاوى  (Complaints)")
    draw_row("إجمالي الشكاوى", str(complaints.get("total_complaints", 0)))
    draw_row("أكثر شكوى",      str(complaints.get("top_complaint", "—")))
    spacer(4)
    if complaints.get("breakdown"):
        draw_table_header(["التصنيف", "العدد", "النسبة"], [230, 80, 80])
        for i, r in enumerate(complaints.get("breakdown") or []):
            draw_table_row([r.get("category","")[:35], _safe_int(r.get("total")), f"{r.get('percent',0)}%"],
                           [230, 80, 80], alt=bool(i % 2))
    spacer()

    # ── 8. تقييمات النجوم ──────────────────────────────────────
    draw_title("8. تقييمات النجوم  (Star Ratings)")
    draw_row("إجمالي التقييمات",  str(_safe_int(stars_ov.get("total_ratings"))))
    draw_row("متوسط النجوم",       f"{_safe_float(stars_ov.get('avg_stars')):.2f} / 5")
    draw_row("نسبة الرضا (4-5)",   f"{stars_ov.get('satisfaction_rate', 0)}%")
    spacer(4)
    draw_table_header(["النجوم", "العدد"], [160, 100])
    for i, r in enumerate(stars_ov.get("distribution") or []):
        stars_str = "★" * int(r.get("stars", 0)) + f"  ({r.get('stars')})"
        draw_table_row([stars_str, _safe_int(r.get("total"))], [160, 100], alt=bool(i % 2))
    spacer()

    # ── 9. التقييمات اليومية ───────────────────────────────────
    draw_title("9. التقييمات اليومية  (Daily Ratings)")
    draw_table_header(["اليوم", "متوسط النجوم", "عدد التقييمات"], [160, 120, 120])
    for i, r in enumerate((stars_daily.get("by_day") or [])[-14:]):
        draw_table_row([r.get("day",""), f"{_safe_float(r.get('avg_stars')):.2f}", _safe_int(r.get("count"))],
                       [160, 120, 120], alt=bool(i % 2))
    spacer()

    # ── 10. آخر التعليقات ──────────────────────────────────────
    recent_with_comment = [x for x in (stars_recent.get("items") or []) if (x.get("comment") or "").strip()]
    if recent_with_comment:
        draw_title("10. آخر التعليقات  (Recent Feedback)")
        for r in recent_with_comment[:8]:
            check_space(30)
            stars_str = "★" * int(r.get("stars") or 0)
            cv.setFont(_pdf_font(bold=True), 8)
            submitted = str(r.get('submitted_at') or '')[:10]
            cv.drawRightString(W - 60, y_state[0], _ar(f"#{r.get('conversation_id')}  {stars_str}  {submitted}"))
            y_state[0] -= 11
            cv.setFont(_pdf_font(), 8)
            comment = (r.get("comment") or "")[:120]
            cv.drawRightString(W - 70, y_state[0], _ar(comment))
            y_state[0] -= 13
        spacer()

    # ── 11. محادثات بتقييم منخفض ───────────────────────────────
    low_items = (low_rated.get("items") or [])
    if low_items:
        draw_title("11. محادثات بتقييم منخفض  (Low-Rated)")
        draw_table_header(["رقم المحادثة", "متوسط النجوم", "عدد التقييمات", "آخر تقييم"], [100, 100, 110, 130])
        for i, c in enumerate(low_items[:20]):
            draw_table_row([
                f"#{_safe_int(c.get('conversation_id'))}",
                f"{_safe_float(c.get('avg_stars')):.1f}",
                _safe_int(c.get("ratings_count")),
                (str(c.get("last_rated_at") or "")[:10]),
            ], [100, 100, 110, 130], alt=bool(i % 2))
        spacer()

    # ── 12. نشاط الموظفين ──────────────────────────────────────
    draw_title("12. نشاط الموظفين  (Employees Activity)")
    emps = sorted(emp.get("employees", []), key=lambda x: int(x.get("active_minutes") or 0), reverse=True)
    draw_table_header(["الاسم", "الدور", "عدد الدخول", "دقائق النشاط", "الحالة"], [140, 70, 80, 90, 80])
    for i, e in enumerate(emps):
        draw_table_row([
            (e.get("name") or "")[:22],
            e.get("role", "")[:10],
            _safe_int(e.get("logins")),
            _safe_int(e.get("active_minutes")),
            e.get("status", ""),
        ], [140, 70, 80, 90, 80], alt=bool(i % 2))

    cv.save()
    buf.seek(0)

    filename = f"full_report_{days}d_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =========================
# ✅ EXPORT: Employees Activity (Excel/PDF) - Admin only
# =========================
@router.get("/export/employees.xlsx")
def export_employees_excel(days: int = 30, _: dict = Depends(_auth_admin)):
    data = _employees_activity_data(days=days)
    employees = data.get("employees", [])

    employees = sorted(employees, key=lambda x: int(x.get("active_minutes") or 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees Activity"
    _ws_rtl(ws)

    total = len(employees)
    active = sum(1 for e in employees if e.get("status") == "نشط")
    low = sum(1 for e in employees if e.get("status") == "قليل النشاط")
    idle = sum(1 for e in employees if e.get("status") == "غير مستخدم")

    ws.append(["ملخص التقرير", ""])
    ws.append(["الفترة (أيام)", data.get("days", days)])
    ws.append(["عدد الموظفين", total])
    ws.append(["نشط", active])
    ws.append(["قليل النشاط", low])
    ws.append(["غير مستخدم", idle])
    ws.append(["تاريخ التوليد (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["", ""])

    header_row = ws.max_row + 1
    ws.append(["#", "رقم الموظف", "الاسم", "الدور", "عدد الدخول", "آخر نشاط", "دقائق النشاط", "الحالة"])

    for idx, emp in enumerate(employees, start=1):
        ws.append(
            [
                idx,
                emp.get("user_id"),
                emp.get("name"),
                emp.get("role"),
                emp.get("logins"),
                _safe_dt(emp.get("last_activity")) or None,
                emp.get("active_minutes"),
                emp.get("status"),
            ]
        )

    _style_header_row(ws, 1)
    _style_header_row(ws, header_row)
    _style_body(ws, 2)

    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = f"A{header_row}:H{ws.max_row}"

    _set_col_formats(ws, {"A": "0", "B": "0", "E": "0", "G": "0"})

    for r in range(header_row + 1, ws.max_row + 1):
        ws[f"F{r}"].number_format = "yyyy-mm-dd hh:mm"

    status_range = f"H{header_row+1}:H{ws.max_row}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"نشط"'], fill=PatternFill("solid", fgColor="DCFCE7")),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"قليل النشاط"'], fill=PatternFill("solid", fgColor="FEF9C3")),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"غير مستخدم"'], fill=PatternFill("solid", fgColor="F3F4F6")),
    )

    _wb_autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    days2 = data.get("days", days)
    filename = f"employees_activity_{days2}d_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/employees.pdf")
def export_employees_pdf(days: int = 30, _: dict = Depends(_auth_admin)):
    data = _employees_activity_data(days=days)
    employees = data.get("employees", [])

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    y = h - 60
    c.setFont(_pdf_font(bold=True), 14)
    c.drawString(50, y, f"Employees Activity Report (Last {data.get('days', days)} Days)")
    y -= 20
    c.setFont(_pdf_font(), 10)
    c.drawString(50, y, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    y -= 25

    def _draw_emp_header(yy):
        c.setFont(_pdf_font(bold=True), 9)
        c.drawString(50,  yy, "ID")
        c.drawString(85,  yy, "Name")
        c.drawString(240, yy, "Role")
        c.drawString(300, yy, "Logins")
        c.drawString(355, yy, "Active(min)")
        c.drawString(435, yy, "Status")
        return yy - 14

    y = _draw_emp_header(y)
    c.setFont(_pdf_font(), 9)

    for emp in employees:
        if y < 70:
            c.showPage()
            y = h - 60
            y = _draw_emp_header(y)
            c.setFont(_pdf_font(), 9)

        c.drawString(50,  y, str(emp.get("user_id") or ""))
        c.drawRightString(235, y, _ar(str(emp.get("name") or "")[:24]))
        c.drawString(240, y, str(emp.get("role") or "")[:8])
        c.drawString(300, y, str(emp.get("logins") or 0))
        c.drawString(355, y, str(emp.get("active_minutes") or 0))
        c.drawRightString(530, y, _ar(str(emp.get("status") or "")))
        y -= 12

    c.save()
    buf.seek(0)

    days2 = data.get("days", days)
    filename = f"employees_activity_{days2}d_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# =========================
# Fix 8: Retrieval Analytics
# =========================

def _retrieval_analytics_data(days: int = 30) -> dict:
    """إحصائيات الـ Retrieval — Qdrant vs PostgreSQL MMR، confidence، hallucination tracking"""
    rows_modes = _q_rows(
        """
        SELECT
            COALESCE(response_mode, 'unknown') AS retrieval_mode,
            COUNT(*) AS total,
            ROUND(AVG(CASE WHEN best_score IS NULL THEN 0 ELSE best_score END), 4) AS avg_confidence
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        GROUP BY retrieval_mode
        ORDER BY total DESC
        """,
        (f"-{days}",),
    )

    avg_conf = _q_scalar(
        """
        SELECT ROUND(AVG(CASE WHEN best_score IS NULL THEN 0 ELSE best_score END), 4)
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        """,
        (f"-{days}",),
    ) or 0.0

    # أسئلة رجعت fallback (hallucination risk)
    fallback_count = _q_scalar(
        """
        SELECT COUNT(*) FROM message
        WHERE message_type = 'assistant'
          AND response_mode LIKE '%%fallback%%'
          AND created_at >= NOW() + (%s || ' days')::interval
        """,
        (f"-{days}",),
    ) or 0

    # أسئلة confidence منخفض (< 0.38)
    low_conf_count = _q_scalar(
        """
        SELECT COUNT(*) FROM message
        WHERE message_type = 'assistant'
          AND best_score IS NOT NULL
          AND best_score < 0.38
          AND created_at >= NOW() + (%s || ' days')::interval
        """,
        (f"-{days}",),
    ) or 0

    total_answered = _q_scalar(
        """
        SELECT COUNT(*) FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        """,
        (f"-{days}",),
    ) or 1

    # أكثر الـ intents
    top_intents = _q_rows(
        """
        SELECT
            COALESCE(intent_pred, 'unknown') AS intent,
            COUNT(*) AS total,
            ROUND(AVG(CASE WHEN best_score IS NULL THEN 0 ELSE best_score END), 4) AS avg_conf
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        GROUP BY intent
        ORDER BY total DESC
        LIMIT 10
        """,
        (f"-{days}",),
    )

    # الأسئلة اللي رجعت fallback (آخر 20)
    fallback_questions = _q_rows(
        """
        SELECT message_text, best_score, response_mode, created_at
        FROM message
        WHERE message_type = 'assistant'
          AND response_mode LIKE '%%fallback%%'
          AND created_at >= NOW() + (%s || ' days')::interval
        ORDER BY created_at DESC
        LIMIT 20
        """,
        (f"-{days}",),
    )

    return {
        "days": days,
        "retrieval_modes": rows_modes,
        "avg_confidence": float(avg_conf),
        "fallback_count": int(fallback_count),
        "low_conf_count": int(low_conf_count),
        "fallback_rate": round(int(fallback_count) / max(1, int(total_answered)), 4),
        "low_conf_rate": round(int(low_conf_count) / max(1, int(total_answered)), 4),
        "top_intents": top_intents,
        "fallback_questions": fallback_questions,
    }


@router.get("/retrieval-analytics")
def retrieval_analytics(
    days: int = Query(30, ge=1, le=365),
    _user=Depends(_auth_stats),
):
    """
    إحصائيات الـ Hybrid Search:
    - كم مرة استخدم hybrid_qdrant vs mmr_pg
    - average confidence
    - fallback rate (hallucination risk)
    - low confidence questions
    - most asked intents
    """
    return _retrieval_analytics_data(days)


@router.get("/hallucination-risk")
def hallucination_risk(
    days: int = Query(30, ge=1, le=365),
    _user=Depends(_auth_stats),
):
    """
    Fix 9: الأسئلة اللي confidence أقل من threshold أو رجعت fallback.
    مهمة لتحسين الـ Knowledge Base.
    """
    rows = _q_rows(
        """
        SELECT
            m.message_text AS question,
            m.best_score AS confidence,
            m.response_mode,
            m.intent_pred AS intent,
            m.created_at
        FROM message m
        WHERE m.message_type = 'assistant'
          AND (
              m.response_mode LIKE '%%fallback%%'
              OR (m.best_score IS NOT NULL AND m.best_score < 0.38)
          )
          AND m.created_at >= NOW() + (%s || ' days')::interval
        ORDER BY m.best_score ASC NULLS FIRST, m.created_at DESC
        LIMIT 50
        """,
        (f"-{days}",),
    )
    return {
        "days": days,
        "count": len(rows),
        "questions": rows,
    }

# ═══════════════════════════════════════════════════════════
# Intent Analytics + Cache Stats — تحسينات جديدة
# ═══════════════════════════════════════════════════════════

@router.get("/intent-breakdown")
def intent_breakdown(days: int = 30, _: dict = Depends(_auth_stats)):
    """
    توزيع الأسئلة حسب الـ intent — مهم لفهم ما يسأل عنه المواطنون.
    يشمل: direct answers، cache hits، off_topic، وكل أنواع الـ RAG.
    """
    rows = _q_rows(
        """
        SELECT
            COALESCE(intent_pred, 'unknown')      AS intent,
            response_mode,
            COUNT(*)                               AS count,
            ROUND(AVG(best_score), 3)              AS avg_confidence,
            ROUND(AVG(
                CASE WHEN response_mode = 'direct_intent' THEN 0
                     WHEN response_mode = 'cache_hit'     THEN 1
                     ELSE NULL END
            ), 3)                                  AS cache_direct_ratio
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        GROUP BY intent_pred, response_mode
        ORDER BY count DESC
        """,
        (f"-{days}",),
    )

    # تجميع حسب intent فقط
    by_intent: dict = {}
    for r in rows:
        intent = r["intent"]
        if intent not in by_intent:
            by_intent[intent] = {
                "intent": intent,
                "total": 0,
                "direct": 0,
                "cached": 0,
                "rag": 0,
                "fallback": 0,
                "off_topic": 0,
                "avg_confidence": [],
            }
        entry = by_intent[intent]
        entry["total"] += r["count"]
        mode = r["response_mode"] or ""
        if mode == "direct_intent":
            entry["direct"] += r["count"]
        elif mode == "cache_hit":
            entry["cached"] += r["count"]
        elif "fallback" in mode:
            entry["fallback"] += r["count"]
        elif mode == "off_topic":
            entry["off_topic"] += r["count"]
        else:
            entry["rag"] += r["count"]
        if r["avg_confidence"] is not None:
            entry["avg_confidence"].append(r["avg_confidence"])

    result = []
    for entry in by_intent.values():
        confs = entry.pop("avg_confidence")
        entry["avg_confidence"] = round(sum(confs) / len(confs), 3) if confs else 0.0
        result.append(entry)

    result.sort(key=lambda x: x["total"], reverse=True)
    return {"days": days, "intents": result}


@router.get("/response-modes-breakdown")
def response_modes_breakdown(days: int = 30, _: dict = Depends(_auth_stats)):
    """
    توزيع طريقة الرد: direct_intent / cache_hit / rag / fallback / off_topic
    يُظهر كفاءة كل مكوّن في النظام.
    """
    rows = _q_rows(
        """
        SELECT
            COALESCE(response_mode, 'unknown') AS mode,
            COUNT(*) AS count,
            ROUND(AVG(best_score), 3) AS avg_confidence
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        GROUP BY response_mode
        ORDER BY count DESC
        """,
        (f"-{days}",),
    )
    total = sum(r["count"] for r in rows)
    for r in rows:
        r["percentage"] = round(r["count"] / max(1, total) * 100, 1)
    return {"days": days, "total": total, "modes": rows}


@router.get("/cache-stats")
def cache_stats(_: dict = Depends(_auth_stats)):
    """
    إحصائيات الـ Answer Cache — hits, misses, hit_rate، وأكثر الأسئلة تكراراً.
    """
    from .answer_cache import get_answer_cache
    cache = get_answer_cache()
    stats = cache.stats()
    top_q = cache.top_questions(n=15)
    return {
        "cache": stats,
        "top_repeated_questions": top_q,
    }


@router.delete("/cache-clear")
def cache_clear(_: dict = Depends(_auth_stats)):
    """
    يمسح الـ Answer Cache — استخدمه بعد تحديث قاعدة المعرفة.
    """
    from .answer_cache import get_answer_cache
    n = get_answer_cache().clear()
    return {"ok": True, "cleared_entries": n}


@router.get("/guardrails-stats")
def guardrails_stats(days: int = 30, _: dict = Depends(_auth_stats)):
    """
    الأسئلة اللي اتصدت كـ off_topic — مهم لمعرفة ما يحاول المواطنون سؤاله.
    """
    rows = _q_rows(
        """
        SELECT
            m.message_text  AS question,
            m.created_at
        FROM message m
        WHERE m.message_type = 'assistant'
          AND m.response_mode = 'off_topic'
          AND m.created_at >= NOW() + (%s || ' days')::interval
        ORDER BY m.created_at DESC
        LIMIT 50
        """,
        (f"-{days}",),
    )
    total_off_topic = _q_scalar(
        """
        SELECT COUNT(*) FROM message
        WHERE message_type = 'assistant'
          AND response_mode = 'off_topic'
          AND created_at >= NOW() + (%s || ' days')::interval
        """,
        (f"-{days}",),
    ) or 0

    total_all = _q_scalar(
        """
        SELECT COUNT(*) FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() + (%s || ' days')::interval
        """,
        (f"-{days}",),
    ) or 1

    return {
        "days": days,
        "total_off_topic": total_off_topic,
        "off_topic_rate": round(total_off_topic / total_all * 100, 2),
        "recent_blocked_questions": rows,
    }

# ══════════════════════════════════════════════════════════════
# ✅ RAG Evaluation Metrics Endpoint (Precision / Recall / F1)
# ══════════════════════════════════════════════════════════════

@router.get("/rag-eval-metrics")
def rag_eval_metrics(
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(_auth_stats),
):
    """
    مقاييس جودة الـ RAG المحسوبة تلقائياً:
    - Precision@K, Recall@K, F1@K, MRR, Hit@K
    - اتجاه يومي لمعرفة هل الجودة تتحسن أم تتراجع
    """
    # ملاحظة: لو جدول rag_eval_log غير موجود بعد نرجع قيم افتراضية
    try:
        avg_row = _q_rows(
            """
            SELECT
                COUNT(*) AS total,
                ROUND(AVG(precision_k)::numeric, 4)  AS avg_precision,
                ROUND(AVG(recall_k)::numeric, 4)     AS avg_recall,
                ROUND(AVG(f1_k)::numeric, 4)         AS avg_f1,
                ROUND(AVG(mrr)::numeric, 4)           AS avg_mrr,
                ROUND(AVG(hit_at_k)::numeric, 4)     AS hit_rate
            FROM rag_eval_log
            WHERE created_at >= NOW() - INTERVAL '%s days'
            """,
            (days,),
        )
        agg = avg_row[0] if avg_row else {}

        daily = _q_rows(
            """
            SELECT
                DATE(created_at) AS day,
                COUNT(*) AS evals,
                ROUND(AVG(precision_k)::numeric, 4) AS precision,
                ROUND(AVG(recall_k)::numeric, 4)    AS recall,
                ROUND(AVG(f1_k)::numeric, 4)        AS f1,
                ROUND(AVG(mrr)::numeric, 4)          AS mrr,
                ROUND(AVG(hit_at_k)::numeric, 4)    AS hit_rate
            FROM rag_eval_log
            WHERE created_at >= NOW() - INTERVAL '%s days'
            GROUP BY DATE(created_at)
            ORDER BY day ASC
            """,
            (days,),
        )

        return {
            "days": days,
            "total_evals": int(agg.get("total") or 0),
            "avg_precision": float(agg.get("avg_precision") or 0),
            "avg_recall":    float(agg.get("avg_recall") or 0),
            "avg_f1":        float(agg.get("avg_f1") or 0),
            "avg_mrr":       float(agg.get("avg_mrr") or 0),
            "hit_rate":      float(agg.get("hit_rate") or 0),
            "daily": daily,
        }
    except Exception as e:
        return {
            "days": days, "total_evals": 0,
            "avg_precision": 0, "avg_recall": 0, "avg_f1": 0,
            "avg_mrr": 0, "hit_rate": 0, "daily": [],
            "note": f"rag_eval_log not ready yet: {e}",
        }

# ══════════════════════════════════════════════════════════════
# ✅ نشاط قاعدة المعرفة — مين أضاف/عدّل/حذف ومتى
# ══════════════════════════════════════════════════════════════

@router.get("/kb-activity")
def kb_activity(
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(50, ge=1, le=200),
    _: dict = Depends(_auth_stats),
):
    """
    سجل تعديلات قاعدة المعرفة — مين أضاف/عدّل/حذف، متى، وماذا غيّر.
    مهم للمدير لمعرفة جودة عمل الموظفين على الـ KB.
    """
    rows = _q_rows(
        """
        SELECT
            c.change_id,
            c.kb_id,
            c.action,
            c.changed_at,
            COALESCE(u.full_name, u.username, 'غير معروف') AS employee_name,
            u.role                                          AS employee_role,
            c.old_question,
            c.new_question,
            LEFT(c.old_answer, 120)                        AS old_answer_preview,
            LEFT(c.new_answer, 120)                        AS new_answer_preview
        FROM kb_changelog c
        LEFT JOIN app_user u ON u.user_id = c.user_id
        WHERE c.changed_at >= NOW() - INTERVAL '%s days'
        ORDER BY c.changed_at DESC
        LIMIT %s
        """,
        (days, limit),
    )
    return {"days": days, "total": len(rows), "changes": rows}


@router.get("/kb-contributors")
def kb_contributors(
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(_auth_stats),
):
    """
    أكثر الموظفين مساهمةً في قاعدة المعرفة — مفيد لتقييم الأداء.
    """
    rows = _q_rows(
        """
        SELECT
            COALESCE(u.full_name, u.username, 'غير معروف') AS employee_name,
            u.role                                          AS employee_role,
            COUNT(*)                                        AS total_changes,
            SUM(CASE WHEN c.action = 'create' THEN 1 ELSE 0 END) AS added,
            SUM(CASE WHEN c.action = 'update' THEN 1 ELSE 0 END) AS updated,
            SUM(CASE WHEN c.action = 'delete' THEN 1 ELSE 0 END) AS deleted,
            MAX(c.changed_at)                               AS last_change
        FROM kb_changelog c
        LEFT JOIN app_user u ON u.user_id = c.user_id
        WHERE c.changed_at >= NOW() - INTERVAL '%s days'
        GROUP BY u.user_id, u.full_name, u.username, u.role
        ORDER BY total_changes DESC
        """,
        (days,),
    )
    return {"days": days, "contributors": rows}


# ══════════════════════════════════════════════════════════════
# ✅ تقرير الموظف الشخصي — ماذا فعل هذا الموظف تحديداً
# ══════════════════════════════════════════════════════════════

@router.get("/employee-report/{user_id}")
def employee_report(
    user_id: int,
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(_auth_stats),
):
    """
    تقرير شامل لموظف بعينه:
    - عدد مرات الدخول وآخر نشاط
    - مساهماته في قاعدة المعرفة (إضافة/تعديل/حذف)
    - سجل آخر تعديلاته
    """
    # معلومات الموظف
    user_row = _q_rows(
        """
        SELECT user_id, full_name, username, role, email,
               last_login, created_at, status
        FROM app_user WHERE user_id = %s
        """,
        (user_id,),
    )
    if not user_row:
        return {"error": "الموظف غير موجود"}

    # نشاط الجلسات
    sessions = _q_rows(
        """
        SELECT
            COUNT(*)                                        AS logins,
            MAX(last_activity)                              AS last_activity,
            SUM(EXTRACT(EPOCH FROM (
                COALESCE(last_activity, created_at) - created_at
            )) / 60)                                        AS total_minutes
        FROM user_session
        WHERE user_id = %s
          AND created_at >= NOW() - INTERVAL '%s days'
        """,
        (user_id, days),
    )
    session_data = sessions[0] if sessions else {}

    # مساهمات KB
    kb_stats = _q_rows(
        """
        SELECT
            COUNT(*)                                        AS total_changes,
            SUM(CASE WHEN action = 'create' THEN 1 ELSE 0 END) AS added,
            SUM(CASE WHEN action = 'update' THEN 1 ELSE 0 END) AS updated,
            SUM(CASE WHEN action = 'delete' THEN 1 ELSE 0 END) AS deleted
        FROM kb_changelog
        WHERE user_id = %s
          AND changed_at >= NOW() - INTERVAL '%s days'
        """,
        (user_id, days),
    )
    kb_data = kb_stats[0] if kb_stats else {}

    # آخر 10 تعديلات
    recent_changes = _q_rows(
        """
        SELECT action, new_question, changed_at
        FROM kb_changelog
        WHERE user_id = %s
          AND changed_at >= NOW() - INTERVAL '%s days'
        ORDER BY changed_at DESC LIMIT 10
        """,
        (user_id, days),
    )

    return {
        "days": days,
        "employee": user_row[0],
        "sessions": {
            "logins":        int(session_data.get("logins") or 0),
            "last_activity": session_data.get("last_activity"),
            "total_minutes": int(float(session_data.get("total_minutes") or 0)),
        },
        "kb_contributions": {
            "total":   int(kb_data.get("total_changes") or 0),
            "added":   int(kb_data.get("added") or 0),
            "updated": int(kb_data.get("updated") or 0),
            "deleted": int(kb_data.get("deleted") or 0),
        },
        "recent_kb_changes": recent_changes,
    }


# ══════════════════════════════════════════════════════════════
# ✅ تقرير المدير الشامل — ملخص مفيد للبلدية
# ══════════════════════════════════════════════════════════════

@router.get("/municipality-summary")
def municipality_summary(
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(_auth_stats),
):
    """
    ملخص تنفيذي للبلدية يشمل:
    - حجم الخدمة: عدد المواطنين المخدومين والمحادثات
    - جودة الخدمة: نسبة الإجابات الصحيحة ومتوسط التقييم
    - أكثر المشاكل المطروحة
    - أكثر الأحياء شكوى
    - فجوات قاعدة المعرفة (أسئلة بلا إجابة)
    - مساهمات الموظفين
    """
    # حجم الخدمة
    total_conversations = _q_scalar(
        "SELECT COUNT(*) FROM conversation WHERE started_at >= NOW() - INTERVAL '%s days'",
        (days,),
    ) or 0

    total_citizens = _q_scalar(
        """
        SELECT COUNT(DISTINCT conversation_id) FROM message
        WHERE message_type = 'user'
          AND created_at >= NOW() - INTERVAL '%s days'
        """,
        (days,),
    ) or 0

    total_questions = _q_scalar(
        """
        SELECT COUNT(*) FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() - INTERVAL '%s days'
        """,
        (days,),
    ) or 0

    # جودة الخدمة
    answer_rate = _q_scalar(
        """
        SELECT ROUND(AVG(CASE WHEN answer_found = 1 THEN 100.0 ELSE 0 END), 1)
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() - INTERVAL '%s days'
        """,
        (days,),
    ) or 0.0

    avg_stars = _q_scalar(
        """
        SELECT ROUND(AVG(rating::numeric), 2)
        FROM feedback
        WHERE feedback_type = 'stars'
          AND submitted_at >= NOW() - INTERVAL '%s days'
        """,
        (days,),
    ) or 0.0

    # أكثر 5 مشاكل
    top_problems = _q_rows(
        """
        SELECT
            COALESCE(intent_pred, 'unknown') AS problem,
            COUNT(*) AS count
        FROM message
        WHERE message_type = 'assistant'
          AND created_at >= NOW() - INTERVAL '%s days'
        GROUP BY intent_pred
        ORDER BY count DESC LIMIT 5
        """,
        (days,),
    )

    # أسئلة بلا إجابة (فجوات الـ KB)
    unanswered_questions = _q_rows(
        """
        SELECT
            message_text AS question,
            COUNT(*) AS times_asked
        FROM message
        WHERE message_type = 'assistant'
          AND (answer_found = 0 OR response_mode LIKE '%%fallback%%')
          AND created_at >= NOW() - INTERVAL '%s days'
          AND message_text IS NOT NULL
        GROUP BY message_text
        ORDER BY times_asked DESC
        LIMIT 10
        """,
        (days,),
    )

    # مساهمات الموظفين في الـ KB
    kb_summary = _q_rows(
        """
        SELECT
            COALESCE(u.full_name, u.username, 'غير معروف') AS name,
            COUNT(*) AS changes
        FROM kb_changelog c
        LEFT JOIN app_user u ON u.user_id = c.user_id
        WHERE c.changed_at >= NOW() - INTERVAL '%s days'
        GROUP BY u.user_id, u.full_name, u.username
        ORDER BY changes DESC LIMIT 5
        """,
        (days,),
    )

    total_kb_entries = _q_scalar(
        "SELECT COUNT(*) FROM knowledge_base WHERE is_active = 1"
    ) or 0

    return {
        "days": days,
        "period_label": f"آخر {days} يوم",
        "service_volume": {
            "total_conversations": int(total_conversations),
            "total_sessions":      int(total_citizens),
            "total_questions":     int(total_questions),
            "daily_avg_questions": round(int(total_questions) / max(1, days), 1),
        },
        "service_quality": {
            "answer_rate_pct": float(answer_rate),
            "avg_stars":       float(avg_stars),
            "kb_entries":      int(total_kb_entries),
        },
        "top_problems": top_problems,
        "knowledge_gaps": {
            "unanswered_count": len(unanswered_questions),
            "top_unanswered":   unanswered_questions,
        },
        "employee_kb_contributions": kb_summary,
        "generated_at": datetime.utcnow().isoformat(),
    }


# ══════════════════════════════════════════════════════════════
# ✅ Export: تقرير قاعدة المعرفة Excel
# ══════════════════════════════════════════════════════════════

@router.get("/export/kb-report.xlsx")
def export_kb_report_excel(
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(_auth_stats),
):
    """
    تقرير Excel لقاعدة المعرفة يشمل:
    - ورقة: ملخص المساهمات (مين أضاف كم)
    - ورقة: سجل التعديلات الكاملة
    - ورقة: الأسئلة بلا إجابة (فجوات تحتاج إضافة)
    """
    contributors_data = _q_rows(
        """
        SELECT
            COALESCE(u.full_name, u.username, 'غير معروف') AS employee_name,
            u.role                                          AS employee_role,
            COUNT(*)                                        AS total_changes,
            SUM(CASE WHEN c.action = 'create' THEN 1 ELSE 0 END) AS added,
            SUM(CASE WHEN c.action = 'update' THEN 1 ELSE 0 END) AS updated,
            SUM(CASE WHEN c.action = 'delete' THEN 1 ELSE 0 END) AS deleted,
            MAX(c.changed_at)                               AS last_change
        FROM kb_changelog c
        LEFT JOIN app_user u ON u.user_id = c.user_id
        WHERE c.changed_at >= NOW() - INTERVAL '%s days'
        GROUP BY u.user_id, u.full_name, u.username, u.role
        ORDER BY total_changes DESC
        """,
        (days,),
    )

    changes_data = _q_rows(
        """
        SELECT
            c.change_id,
            c.kb_id,
            c.action,
            COALESCE(u.full_name, u.username, 'غير معروف') AS employee_name,
            COALESCE(c.new_question, c.old_question)        AS question,
            LEFT(COALESCE(c.new_answer, c.old_answer), 200) AS answer_preview,
            c.changed_at
        FROM kb_changelog c
        LEFT JOIN app_user u ON u.user_id = c.user_id
        WHERE c.changed_at >= NOW() - INTERVAL '%s days'
        ORDER BY c.changed_at DESC
        LIMIT 500
        """,
        (days,),
    )

    gaps_data = _q_rows(
        """
        SELECT
            message_text AS question,
            COUNT(*)     AS times_asked,
            MAX(created_at) AS last_asked
        FROM message
        WHERE message_type = 'assistant'
          AND (answer_found = 0 OR response_mode LIKE '%%fallback%%')
          AND created_at >= NOW() - INTERVAL '%s days'
          AND message_text IS NOT NULL
        GROUP BY message_text
        ORDER BY times_asked DESC
        LIMIT 100
        """,
        (days,),
    )

    wb = Workbook()

    # ── Sheet 1: مساهمات الموظفين ─────────────────────────────
    ws1 = wb.active
    ws1.title = "مساهمات الموظفين"
    _ws_rtl(ws1)
    ws1.append(["الموظف", "الدور", "إجمالي", "أضاف", "عدّل", "حذف", "آخر تعديل"])
    for r in contributors_data:
        ws1.append([
            r.get("employee_name"), r.get("employee_role"),
            _safe_int(r.get("total_changes")),
            _safe_int(r.get("added")),
            _safe_int(r.get("updated")),
            _safe_int(r.get("deleted")),
            _safe_dt(r.get("last_change")) or "",
        ])
    _style_header_row(ws1, 1)
    _style_body(ws1, 2)
    _freeze_and_filter(ws1, "A2")
    for r in range(2, ws1.max_row + 1):
        ws1[f"G{r}"].number_format = "yyyy-mm-dd hh:mm"
    _wb_autofit(ws1)

    # ── Sheet 2: سجل التعديلات ────────────────────────────────
    ws2 = wb.create_sheet("سجل التعديلات")
    _ws_rtl(ws2)
    ws2.append(["#", "رقم KB", "الإجراء", "الموظف", "السؤال", "معاينة الإجابة", "التاريخ"])
    for idx, r in enumerate(changes_data, 1):
        action_ar = {"create": "إضافة", "update": "تعديل", "delete": "حذف"}.get(
            str(r.get("action", "")), r.get("action", "")
        )
        ws2.append([
            idx,
            _safe_int(r.get("kb_id")),
            action_ar,
            r.get("employee_name") or "",
            (r.get("question") or "")[:100],
            (r.get("answer_preview") or "")[:200],
            _safe_dt(r.get("changed_at")) or "",
        ])
    _style_header_row(ws2, 1)
    _style_body(ws2, 2)
    _freeze_and_filter(ws2, "A2")
    for r in range(2, ws2.max_row + 1):
        ws2[f"G{r}"].number_format = "yyyy-mm-dd hh:mm"
    _wb_autofit(ws2)

    # ── Sheet 3: فجوات قاعدة المعرفة ─────────────────────────
    ws3 = wb.create_sheet("فجوات المعرفة")
    _ws_rtl(ws3)
    ws3.append(["السؤال (بلا إجابة)", "عدد مرات السؤال", "آخر مرة"])
    for r in gaps_data:
        ws3.append([
            (r.get("question") or "")[:200],
            _safe_int(r.get("times_asked")),
            _safe_dt(r.get("last_asked")) or "",
        ])
    _style_header_row(ws3, 1)
    _style_body(ws3, 2)
    _freeze_and_filter(ws3, "A2")
    # لوّن الأسئلة الأكثر تكراراً بالأحمر الفاتح
    from openpyxl.formatting.rule import ColorScaleRule
    if ws3.max_row > 1:
        ws3.conditional_formatting.add(
            f"B2:B{ws3.max_row}",
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max",   end_color="FF9999",
            ),
        )
    for r in range(2, ws3.max_row + 1):
        ws3[f"C{r}"].number_format = "yyyy-mm-dd"
    _wb_autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kb_report_{days}d_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
